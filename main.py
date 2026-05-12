"""CLI for the data layer.

Usage:
  python main.py download --start 2025-12-18 --end 2025-12-19 --sources binance_futures
  python main.py download                          # use period from config.yaml
  python main.py status                            # show cached ranges per source

Sources: binance_futures, binance_spot, bybit_linear, binance_aggtrades, all
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd
import yaml

from data import (
    binance_aggtrades,
    binance_futures,
    binance_spot,
    bybit,
    storage,
)
from analysis import outcomes as outcomes_mod
from analysis import series as series_mod
from analysis import features as features_mod
from analysis import backtest as backtest_mod
from analysis import charts as charts_mod

ALL_SOURCES = ["binance_futures", "binance_spot", "bybit_linear", "binance_aggtrades"]


def load_config(path: str = "config.yaml") -> dict:
    # Mirror tui._load_cfg: cwd -> next to .exe -> bundled in .exe (fallback).
    candidates = [Path(path)]
    if getattr(sys, "frozen", False):
        candidates.append(Path(sys.executable).parent / path)
        meipass = getattr(sys, "_MEIPASS", None)
        if meipass:
            candidates.append(Path(meipass) / path)
    for p in candidates:
        if p.exists():
            with open(p, encoding="utf-8") as f:
                return yaml.safe_load(f)
    raise FileNotFoundError(
        f"config.yaml не найден. Искал: {[str(c) for c in candidates]}"
    )


def parse_ts(s: str) -> pd.Timestamp:
    return pd.Timestamp(s, tz="UTC")


def _klines_download(
    source: str,
    fetcher,
    symbol: str,
    interval: str,
    start: pd.Timestamp,
    end: pd.Timestamp,
    cache_dir: str,
) -> None:
    path = storage.cache_path(cache_dir, source, symbol, interval)
    step = pd.Timedelta(interval)
    cached = storage.load(path)

    # Plan: fetch the missing pieces both before cached_min and after cached_max.
    if cached is None or cached.empty:
        windows = [(start, end)]
        cached_min = cached_max = None
    else:
        cached_min = cached["open_time"].min()
        cached_max = cached["open_time"].max()
        windows = []
        if start < cached_min:
            windows.append((start, cached_min - step))         # backfill earlier
        if end > cached_max:
            windows.append((cached_max + step, end))           # extend forward

    if not windows:
        print(f"[{source}] cache up-to-date ({cached_min} -> {cached_max})")
        return

    for w_start, w_end in windows:
        if w_start > w_end:
            continue
        print(f"[{source}] fetching {w_start} -> {w_end}")
        df = fetcher(symbol, interval, w_start, w_end)
        storage.merge_save(df, path, key="open_time")

    merged = storage.load(path)
    rows = 0 if merged is None else len(merged)
    print(f"[{source}] cached rows: {rows} -> {path}")


def _aggtrades_download(
    symbol: str,
    start: pd.Timestamp,
    end: pd.Timestamp,
    cache_dir: str,
) -> None:
    source = "binance_aggtrades_5m"
    days = list(binance_aggtrades.daterange(start, end))
    total = len(days)
    done = skipped = failed = 0
    for i, day in enumerate(days, 1):
        out = storage.daily_cache_path(cache_dir, source, symbol, day.isoformat())
        if out.exists():
            skipped += 1
            continue
        print(f"[{source}] {i}/{total}  {day}  downloading...", flush=True)
        try:
            agg = binance_aggtrades.fetch_day_5m(symbol, day)
        except Exception as e:  # noqa: BLE001
            failed += 1
            print(f"[{source}] {day} FAILED: {e}")
            continue
        storage.save(agg, out)
        done += 1
        print(f"[{source}] {i}/{total}  {day}  ok ({len(agg)} buckets)", flush=True)
    print(f"[{source}] summary: done={done} skipped(cached)={skipped} failed={failed}")


def cmd_download(args, cfg) -> None:
    symbol = cfg["symbol"]
    interval = cfg["interval"]
    cache_dir = cfg["cache_dir"]

    start = parse_ts(args.start) if args.start else parse_ts(cfg["period"]["start"])
    end = parse_ts(args.end) if args.end else parse_ts(cfg["period"]["end"])

    if args.sources == ["all"]:
        sources = ALL_SOURCES
    else:
        sources = args.sources

    if "binance_futures" in sources:
        _klines_download("binance_futures", binance_futures.fetch_klines, symbol, interval, start, end, cache_dir)
    if "binance_spot" in sources:
        _klines_download("binance_spot", binance_spot.fetch_klines, symbol, interval, start, end, cache_dir)
    if "bybit_linear" in sources:
        _klines_download("bybit_linear", bybit.fetch_klines, symbol, interval, start, end, cache_dir)
    if "binance_aggtrades" in sources:
        _aggtrades_download(symbol, start, end, cache_dir)


def _concat_aggtrades(cache_dir: str, symbol: str) -> pd.DataFrame | None:
    agg_dir = Path(cache_dir) / "binance_aggtrades_5m" / symbol
    if not agg_dir.exists():
        return None
    files = sorted(agg_dir.glob("*.parquet"))
    if not files:
        return None
    parts = [pd.read_parquet(f) for f in files]
    return pd.concat(parts, ignore_index=True).sort_values("bucket").reset_index(drop=True)


def _write_excel(df: pd.DataFrame, path: Path) -> None:
    """Excel can't store tz-aware timestamps; strip the tz but keep each
    column's LOCAL clock time (UTC stays UTC, Moscow stays Moscow)."""
    df = df.copy()
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
    df.to_excel(path, index=False, engine="openpyxl")


def cmd_export(args, cfg) -> None:
    symbol = cfg["symbol"]
    interval = cfg["interval"]
    cache_dir = cfg["cache_dir"]
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    fmt = args.format  # 'xlsx' or 'csv'

    def _dump(name: str, df: pd.DataFrame | None) -> None:
        if df is None or df.empty:
            print(f"[export] {name}: empty, skipped")
            return
        ext = "xlsx" if fmt == "xlsx" else "csv"
        out = out_dir / f"{name}.{ext}"
        if fmt == "xlsx":
            _write_excel(df, out)
        else:
            df.to_csv(out, index=False, sep=args.csv_sep)
        print(f"[export] {name}: {len(df)} rows -> {out}")

    for src in ["binance_futures", "binance_spot", "bybit_linear"]:
        p = storage.cache_path(cache_dir, src, symbol, interval)
        _dump(f"{src}_{symbol}_{interval}", storage.load(p))

    _dump(f"binance_aggtrades_5m_{symbol}", _concat_aggtrades(cache_dir, symbol))

    # analysis artefacts (one set per source)
    analysis_root = Path(cache_dir) / "analysis"
    if analysis_root.exists():
        for src_dir in sorted(analysis_root.iterdir()):
            if not src_dir.is_dir():
                continue
            for kind in ("outcomes", "series", "signals", "long_series_context"):
                p = src_dir / f"{kind}.parquet"
                if p.exists():
                    _dump(f"analysis_{src_dir.name}_{kind}", pd.read_parquet(p))


def cmd_analyze(args, cfg) -> None:
    symbol = cfg["symbol"]
    interval = cfg["interval"]
    cache_dir = cfg["cache_dir"]
    src = args.source

    p = storage.cache_path(cache_dir, src, symbol, interval)
    klines = storage.load(p)
    if klines is None or klines.empty:
        print(f"[analyze] no cached klines for {src}; run `download` first")
        return

    print(f"[analyze] source={src} rule={args.outcome_rule} max_attempts={args.max_attempts}")
    outcomes = outcomes_mod.reconstruct(klines, rule=args.outcome_rule)
    series = series_mod.detect_series(outcomes)
    signals = series_mod.generate_signals(
        outcomes,
        max_attempts=args.max_attempts,
        long_threshold=args.long_threshold,
    )

    # Compute feature columns for each signal and a separate long-series table.
    signals_with_features = features_mod.add_features(signals, klines)
    long_ctx = features_mod.long_series_context(series, klines, threshold=args.long_threshold)

    out_dir = Path(cache_dir) / "analysis" / src
    out_dir.mkdir(parents=True, exist_ok=True)
    storage.save(outcomes, out_dir / "outcomes.parquet")
    storage.save(series, out_dir / "series.parquet")
    storage.save(signals_with_features, out_dir / "signals.parquet")
    if not long_ctx.empty:
        storage.save(long_ctx, out_dir / "long_series_context.parquet")

    n_up = int((outcomes["outcome"] == "UP").sum())
    n_down = int((outcomes["outcome"] == "DOWN").sum())
    stats = series_mod.summarize(series, signals)

    print(f"  Исходов     : {len(outcomes)}  UP={n_up}  DOWN={n_down}  (доля UP {n_up/len(outcomes):.1%})")
    print(f"  Серий       : всего={stats['n_series']}  макс_длина={stats['max_series_len']}  "
          f"≥7: {stats['n_series_ge7']}  ≥10: {stats['n_series_ge10']}")
    print(f"  Сигналов    : всего={stats['n_signals']}  побед={stats['n_wins']}  "
          f"проигрышей={stats['n_losses']}  win_rate={stats['win_rate']:.1%}")
    if stats["n_signals"]:
        print(f"  Доля длинных серий в сигналах (≥7): "
              f"{int(signals['is_long_series'].sum())} ({signals['is_long_series'].mean():.1%})")

    # Сравнение wins vs длинные-серии-losses по ключевым признакам.
    _print_feature_compare(signals_with_features)
    print(f"  Сохранено  -> {out_dir}")


def _print_feature_compare(sigs: pd.DataFrame) -> None:
    wins = sigs[sigs["result"] == "win"]
    longs = sigs[sigs["is_long_series"] == True]  # noqa: E712
    if wins.empty or longs.empty:
        return
    cols = [
        ("volume_ratio",            "объём 3 свечей / 3x среднее за час"),
        ("range_expansion",         "расширение диапазона (3 свечи / 60м)"),
        ("taker_with_series_ratio", "taker в направлении серии"),
        ("body_to_range_3bar_avg",  "среднее body/range (3 свечи)"),
        ("close_position_3bar_avg", "среднее положение close в свече"),
        ("breakout_60m_with_series","пробой 60м в направлении серии"),
        ("breakout_3h_with_series", "пробой 3ч в направлении серии"),
        ("breakout_24h_with_series","пробой 24ч в направлении серии"),
        ("distance_from_ema20_pct", "расстояние от EMA20, %"),
        ("distance_from_vwap_pct",  "расстояние от VWAP, %"),
    ]
    print(f"\n  Сравнение по признакам  (победы n={len(wins)}  vs  длинные серии n={len(longs)}):")
    print(f"    {'признак':40s}  {'победы':>10s}  {'длинные':>10s}  {'разница':>10s}")
    for col, label in cols:
        if col not in sigs.columns:
            continue
        w = wins[col].astype(float).mean()
        l = longs[col].astype(float).mean()
        if pd.isna(w) or pd.isna(l):
            continue
        delta = l - w
        print(f"    {label:40s}  {w:>10.3f}  {l:>10.3f}  {delta:>+10.3f}")


def cmd_backtest(args, cfg) -> None:
    symbol = cfg["symbol"]
    interval = cfg["interval"]
    cache_dir = cfg["cache_dir"]

    src = args.source
    sigs_path = Path(cache_dir) / "analysis" / src / "signals.parquet"
    series_path = Path(cache_dir) / "analysis" / src / "series.parquet"
    if not sigs_path.exists():
        print(f"[backtest] нет сигналов для {src}. Сначала прогон 'Анализ'.")
        return
    sigs = pd.read_parquet(sigs_path)
    series_df = pd.read_parquet(series_path) if series_path.exists() else pd.DataFrame()

    # Multi-payout sweep: principal payout первым, остальные считаются
    # отдельно и попадают на лист "Sweep_по_payout".
    payouts = [float(p.strip()) for p in str(args.payout).split(",") if p.strip()]
    if not payouts:
        payouts = [1.8]
    primary_payout = payouts[0]

    cfg_bt = backtest_mod.BacktestConfig(
        payout=primary_payout,
        base_stake=args.base_stake,
        max_attempts=args.max_attempts,
        stake_mode=args.stake_mode,
    )
    stakes = backtest_mod._stakes(cfg_bt)
    cumloss = sum(stakes)

    print(f"\n[backtest] источник={src}")
    print(f"  Параметры стратегии:")
    print(f"    выплата (payout)   : {primary_payout}x  "
          + (f"(плюс sweep: {payouts[1:]})" if len(payouts) > 1 else ""))
    print(f"    догон              : {cfg_bt.max_attempts} ставки")
    print(f"    ставки             : {stakes}  (режим: {cfg_bt.stake_mode})")
    print(f"    проигрыш всей цепи : -{cumloss}")
    print(f"    train_end          : {args.train_end}")

    results = backtest_mod.run_full_backtest(sigs, cfg_bt, train_end=args.train_end)

    out_dir = Path(cache_dir) / "analysis" / src
    storage.save(results, out_dir / "backtest_results.parquet")

    # Sweep по другим payout-ам (компактная таблица по 'весь_период')
    sweep_rows: list[dict] = []
    if len(payouts) > 1:
        for p in payouts:
            cfg_p = backtest_mod.BacktestConfig(
                payout=p, base_stake=cfg_bt.base_stake,
                max_attempts=cfg_bt.max_attempts, stake_mode=cfg_bt.stake_mode,
            )
            r = backtest_mod.run_full_backtest(sigs, cfg_p, train_end=args.train_end)
            r_all = r[r["период"] == "весь_период"][[
                "стратегия", "сигналов_всего", "пропущено", "осталось",
                "win_rate_%", "PnL_всего", "PnL_на_сигнал", "max_drawdown",
            ]].copy()
            r_all.insert(0, "payout", p)
            sweep_rows.append(r_all)
    sweep_df = pd.concat(sweep_rows, ignore_index=True) if sweep_rows else pd.DataFrame()

    # Печатаем компактные таблицы по периодам
    print("\n  Результаты (весь период):")
    _print_backtest_table(results[results["период"] == "весь_период"])
    print("\n  Walk-forward train (до train_end):")
    _print_backtest_table(results[results["период"] == "train"])
    print("\n  Walk-forward test (после train_end):")
    _print_backtest_table(results[results["период"] == "test"])

    # Графики (7 PNG из ТЗ файл 4)
    charts_dir = Path(args.out) / "charts"
    sigs_with_pnl = backtest_mod.compute_pnl(
        backtest_mod.recompute_outcome(sigs, cfg_bt), cfg_bt
    )
    chart_files = charts_mod.make_all(sigs_with_pnl, series_df, charts_dir, long_threshold=7)
    if chart_files:
        print(f"\n  Графиков сделано: {len(chart_files)}  -> {charts_dir}")

    # Excel-отчёт
    out_xlsx = Path(args.out) / f"backtest_filters_report_{src}.xlsx"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    _write_backtest_excel(results, out_xlsx, cfg_bt, stakes, args.train_end,
                          sweep_df, chart_files)
    print(f"\n  Отчёт   -> {out_xlsx}")
    print(f"  Кэш     -> {out_dir / 'backtest_results.parquet'}")


def _print_backtest_table(df: pd.DataFrame) -> None:
    if df.empty:
        print("    (пусто)")
        return
    print(f"    {'стратегия':30s} {'сигн':>5s} {'проп':>5s} {'осн':>5s} "
          f"{'побед':>6s} {'пр-ш':>5s} {'win_%':>6s} {'PnL':>9s} {'PnL/сигн':>10s} {'DD':>9s}")
    # Сортируем по PnL по убыванию, чтобы лучшие сверху
    for _, r in df.sort_values("PnL_всего", ascending=False).iterrows():
        print(f"    {r['стратегия']:30s} {int(r['сигналов_всего']):>5d} "
              f"{int(r['пропущено']):>5d} {int(r['осталось']):>5d} "
              f"{int(r['побед']):>6d} {int(r['проигрышей']):>5d} "
              f"{r['win_rate_%']:>6.1f} {r['PnL_всего']:>+9.2f} "
              f"{r['PnL_на_сигнал']:>+10.3f} {r['max_drawdown']:>+9.2f}")


def _write_backtest_excel(
    results: pd.DataFrame,
    path: Path,
    cfg,
    stakes,
    train_end: str,
    sweep_df: pd.DataFrame | None = None,
    chart_files: list | None = None,
) -> None:
    """Многолистовой xlsx-отчёт по бэктесту для заказчика."""
    chart_files = chart_files or []
    sweep_df = sweep_df if sweep_df is not None else pd.DataFrame()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Лист 1: Выводы — авто-генерация по фактам из results
        conclusions = _build_conclusions(results, cfg, stakes)
        conclusions.to_excel(writer, sheet_name="Выводы", index=False)

        # Лист 2: Сводка — все стратегии × все периоды
        summary = results[[
            "стратегия", "период", "сигналов_всего", "пропущено", "осталось",
            "побед", "проигрышей", "win_rate_%", "PnL_всего", "PnL_на_сигнал", "max_drawdown",
        ]].copy()
        summary.to_excel(writer, sheet_name="Сводка", index=False)

        # Лист 3: Описание фильтров
        descriptions = (
            results[["стратегия_код", "стратегия", "описание"]]
            .drop_duplicates(subset=["стратегия_код"])
            .reset_index(drop=True)
        )
        descriptions.to_excel(writer, sheet_name="Описание_фильтров", index=False)

        # Лист 4: Параметры стратегии и интерпретация
        params = pd.DataFrame([
            ["Выплата (payout)",            f"{cfg.payout}x"],
            ["Догон (макс. ставок)",        cfg.max_attempts],
            ["Базовая ставка",              cfg.base_stake],
            ["Режим ставок",                cfg.stake_mode],
            ["Ставки по шагам",             ", ".join(f"{s:g}" for s in stakes)],
            ["Сумма проигрыша всей цепи",   -sum(stakes)],
            ["Граница train/test",          train_end],
            ["",                            ""],
            ["Что показывает Сводка",       "Одна строка = одна стратегия × один период."],
            ["Колонка 'пропущено'",         "Сколько сигналов отбросил фильтр."],
            ["Колонка 'осталось'",          "Сигналы, по которым ставили."],
            ["Колонка 'win_rate_%'",        "Процент побед среди оставшихся сигналов."],
            ["Колонка 'PnL_всего'",         "Сумма прибыли/убытка в единицах базовой ставки."],
            ["Колонка 'PnL_на_сигнал'",     "PnL_всего / число оставшихся сигналов."],
            ["Колонка 'max_drawdown'",      "Максимальная просадка эквити (отрицательное число)."],
        ], columns=["параметр", "значение"])
        params.to_excel(writer, sheet_name="Параметры", index=False)

        # Лист 5: Sweep по разным payout (если был запрошен)
        if not sweep_df.empty:
            sweep_df.to_excel(writer, sheet_name="Sweep_по_payout", index=False)

        # Лист 6: Графики (вставляем PNG прямо в Excel)
        if chart_files:
            from openpyxl.drawing.image import Image as XLImage
            ws = writer.book.create_sheet("Графики")
            ws["A1"] = "7 графиков по ТЗ заказчика (File 4). См. также папку export/charts/"
            row = 3
            for fp in chart_files:
                try:
                    img = XLImage(str(fp))
                    # Уменьшаем размеры до приличных в Excel
                    img.width = 720
                    img.height = 400
                    ws.add_image(img, f"A{row}")
                    row += 22
                except Exception as e:  # noqa: BLE001
                    ws.cell(row=row, column=1).value = f"[не удалось вставить {fp.name}: {e}]"
                    row += 2


def _build_conclusions(results: pd.DataFrame, cfg, stakes) -> pd.DataFrame:
    """Авто-выводы на русском, заполнены реальными числами из результатов."""
    all_p = results[results["период"] == "весь_период"].copy()
    train = results[results["период"] == "train"].copy()
    test = results[results["период"] == "test"].copy()

    baseline = all_p[all_p["стратегия_код"] == "baseline"].iloc[0] if "baseline" in all_p["стратегия_код"].values else None
    # "Лучшую" ищем только среди стратегий, оставивших содержательное число
    # сигналов: иначе фильтр, отрезающий вообще всё, всегда побеждает с PnL=0.
    MIN_KEPT = 30
    viable = all_p[all_p["осталось"] >= MIN_KEPT]
    best_all = viable.sort_values("PnL_всего", ascending=False).iloc[0] if not viable.empty else None
    viable_tr = train[train["осталось"] >= MIN_KEPT]
    best_train = viable_tr.sort_values("PnL_всего", ascending=False).iloc[0] if not viable_tr.empty else None

    tz_codes = [c for c in all_p["стратегия_код"].unique() if c.startswith("tz_")]
    inv_codes = [c for c in all_p["стратегия_код"].unique() if c.startswith("inv_")]
    tz_pnl = all_p[all_p["стратегия_код"].isin(tz_codes)]["PnL_всего"].mean()
    inv_pnl = all_p[all_p["стратегия_код"].isin(inv_codes)]["PnL_всего"].mean()

    lines: list[tuple[str, str]] = []
    lines.append(("ВЫВОДЫ ПО БЭКТЕСТУ", ""))
    lines.append(("Параметры стратегии",
                  f"мартингейл против серии, payout {cfg.payout}x, "
                  f"догон {cfg.max_attempts}, ставки {stakes} ({cfg.stake_mode})"))
    lines.append(("", ""))

    if baseline is not None:
        lines.append(("1. Базовая стратегия без фильтров",
                      f"win_rate={baseline['win_rate_%']}%  PnL={baseline['PnL_всего']:+.1f}  "
                      f"DD={baseline['max_drawdown']:+.1f}  "
                      f"({int(baseline['сигналов_всего'])} сигналов)"))
        if baseline["PnL_всего"] < 0:
            lines.append(("   Интерпретация",
                          "Стратегия без фильтров в минусе. Высокий win_rate "
                          "не компенсирует -7 за полный проигрыш цепи (1-2-4)."))
        elif baseline["PnL_всего"] > 0:
            lines.append(("   Интерпретация", "Стратегия даже без фильтров прибыльна."))
    lines.append(("", ""))

    if best_all is not None:
        lines.append(("2. Лучшая стратегия (весь период)",
                      f"{best_all['стратегия']}: PnL={best_all['PnL_всего']:+.1f}, "
                      f"win_rate={best_all['win_rate_%']}%, "
                      f"осталось сигналов {int(best_all['осталось'])} из "
                      f"{int(best_all['сигналов_всего'])}"))
    if best_train is not None:
        lines.append(("3. Walk-forward проверка",
                      f"Лучшая на train (≥{MIN_KEPT} сигналов): '{best_train['стратегия']}' "
                      f"PnL={best_train['PnL_всего']:+.1f}."))
        same_on_test = test[test["стратегия_код"] == best_train["стратегия_код"]]
        if not same_on_test.empty:
            r = same_on_test.iloc[0]
            verdict = "результат переносится" if r["PnL_всего"] >= 0 else "результат на test слабее"
            lines.append(("   Та же стратегия на test",
                          f"PnL={r['PnL_всего']:+.1f}, win_rate={r['win_rate_%']}% — {verdict}."))
    lines.append(("", ""))

    if tz_codes and inv_codes:
        lines.append(("4. Гипотеза заказчика vs альтернатива",
                      f"Средний PnL по 5 фильтрам из ТЗ:        {tz_pnl:+.1f}"))
        lines.append(("", f"Средний PnL по 5 инверт-фильтрам:        {inv_pnl:+.1f}"))
        if inv_pnl > tz_pnl:
            lines.append(("   Вывод",
                          "Инверт-фильтры дают больший PnL в среднем. Это значит: "
                          "длинные серии чаще возникают в ТИХИХ условиях (нет пробоя, "
                          "низкий объём, баланс taker), а не в импульсе. "
                          "Гипотеза 'избегать сигналов на пробое' данными не подтверждается."))
        else:
            lines.append(("   Вывод",
                          "Фильтры ТЗ выигрывают у инверт-фильтров. "
                          "Гипотеза заказчика частично подтверждается."))
    lines.append(("", ""))

    n_signals = int(baseline["сигналов_всего"]) if baseline is not None else 0
    lines.append(("5. Что это значит на практике", ""))
    lines.append(("",
                  f"На {n_signals} сигналах за период даже лучший фильтр "
                  f"оставляет нас близко к нулю на test. Чтобы стратегия "
                  f"стабильно выходила в плюс, нужно одно из:"))
    lines.append(("   а)", "Полу­чать payout выше 1.8 (попробуй sweep 1.85 / 1.9 / 2.0)."))
    lines.append(("   б)", "Сократить размер потери цепи: меньше догонов или фикс. ставка."))
    lines.append(("   в)", "Добавить более жёсткий фильтр — например, по часам МСК (см. графики)."))
    lines.append(("", ""))

    lines.append(("6. Как читать остальные листы", ""))
    lines.append(("   Сводка",          "Полная таблица: 11+ стратегий × 3 периода."))
    lines.append(("   Описание_фильтров","Что каждый фильтр делает, по-русски."))
    lines.append(("   Параметры",       "Параметры стратегии и интерпретация колонок."))
    lines.append(("   Графики",         "7 диаграмм из ТЗ (по часам МСК, объёму, пробоям, длинам серий)."))

    return pd.DataFrame(lines, columns=["раздел", "значение"])


def cmd_status(_args, cfg) -> None:
    symbol = cfg["symbol"]
    interval = cfg["interval"]
    cache_dir = cfg["cache_dir"]

    for src in ["binance_futures", "binance_spot", "bybit_linear"]:
        p = storage.cache_path(cache_dir, src, symbol, interval)
        df = storage.load(p)
        if df is None or df.empty:
            print(f"[{src}] empty")
            continue
        print(f"[{src}] rows={len(df)}  {df['open_time'].min()} -> {df['open_time'].max()}")

    agg_dir = Path(cache_dir) / "binance_aggtrades_5m" / symbol
    if agg_dir.exists():
        files = sorted(agg_dir.glob("*.parquet"))
        print(f"[binance_aggtrades_5m] daily files: {len(files)}  ({files[0].stem if files else '-'} .. {files[-1].stem if files else '-'})")
    else:
        print("[binance_aggtrades_5m] empty")


def main() -> None:
    # No arguments at all -> launch the interactive TUI. This is the
    # entry-point most non-developer users (and the PyInstaller .exe) will hit.
    if len(sys.argv) == 1:
        from tui import run as tui_run
        tui_run()
        return

    cfg = load_config()
    parser = argparse.ArgumentParser(prog="b_parser")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_dl = sub.add_parser("download", help="Download/refresh data")
    p_dl.add_argument("--start", help="UTC start (e.g. 2025-12-18 or 2025-12-18T00:00:00Z)")
    p_dl.add_argument("--end", help="UTC end")
    p_dl.add_argument(
        "--sources",
        nargs="+",
        default=["all"],
        choices=ALL_SOURCES + ["all"],
        help="One or more sources to download",
    )
    p_dl.set_defaults(func=cmd_download)

    p_st = sub.add_parser("status", help="Show cached ranges per source")
    p_st.set_defaults(func=cmd_status)

    p_ex = sub.add_parser("export", help="Export cache to Excel/CSV")
    p_ex.add_argument("--out", default="export", help="Output directory (default: ./export)")
    p_ex.add_argument(
        "--format", default="xlsx", choices=["xlsx", "csv"],
        help="Output format (default: xlsx)",
    )
    p_ex.add_argument(
        "--csv-sep", default=";",
        help="CSV separator (default: ';' for RU Excel; use ',' for international)",
    )
    p_ex.set_defaults(func=cmd_export)

    p_an = sub.add_parser("analyze", help="Reconstruct UP/DOWN, detect series, label martingale signals")
    p_an.add_argument(
        "--source", default="binance_spot",
        choices=["binance_spot", "binance_futures", "bybit_linear"],
        help="Klines source to use (default: binance_spot, closest to Polymarket resolution)",
    )
    p_an.add_argument(
        "--outcome-rule", default="candle", choices=["candle", "close_diff"],
        help="'candle' = close>open in same bar; 'close_diff' = close[t]>close[t-1]",
    )
    p_an.add_argument(
        "--max-attempts", type=int, default=3,
        help="Number of martingale bets after the 3-in-a-row trigger (default: 3)",
    )
    p_an.add_argument(
        "--long-threshold", type=int, default=7,
        help="Series length considered 'long' to flag (default: 7)",
    )
    p_an.set_defaults(func=cmd_analyze)

    p_bt = sub.add_parser("backtest", help="Прогнать бэктест мартингейла с фильтрами")
    p_bt.add_argument("--source", default="binance_spot",
                      choices=["binance_spot", "binance_futures", "bybit_linear"])
    p_bt.add_argument("--payout", default="1.8",
                      help="Множитель выплаты (одно число или через запятую для sweep: 1.8,1.9,2.0)")
    p_bt.add_argument("--base-stake", type=float, default=1.0)
    p_bt.add_argument("--max-attempts", type=int, default=3,
                      help="Сколько ставок-догонов после тройки (по дефолту 3)")
    p_bt.add_argument("--stake-mode", default="doubling",
                      choices=["doubling", "fixed"])
    p_bt.add_argument("--train-end", default="2026-03-31T23:59:59Z",
                      help="Граница train/test для walk-forward")
    p_bt.add_argument("--out", default="export",
                      help="Куда положить Excel-отчёт")
    p_bt.set_defaults(func=cmd_backtest)

    def _cmd_tui(_args, _cfg):
        from tui import run as tui_run
        tui_run()

    p_tui = sub.add_parser("tui", help="Launch interactive menu")
    p_tui.set_defaults(func=_cmd_tui)

    args = parser.parse_args()
    args.func(args, cfg)


if __name__ == "__main__":
    main()
